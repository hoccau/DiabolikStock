#!/usr/bin/python3
# -*- coding: utf-8 -*-

"""
Export malle as xlsx file
"""

from openpyxl import Workbook, styles
from PyQt5.QtCore import QDate
import logging

def write_file(query, type_, filename):
    malles = query.get_malles_by_type(type_)

    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['B'].width = 28
    ws.row_dimensions[9].font = styles.Font(bold=True)
    ws.row_dimensions[2].font = styles.Font(bold=True, size=14)
    ws['B2'] = "Type :"
    ws['C2'] = type_
    ws['B6'] = "Références des caisses : "
    ws['C6'] = ", ".join(malles)
    ws['B7'] = "Nombre total de caisses : "
    ws['C7'] = len(malles)
    ws['B9'] = "Désignation produit"
    ws['C9'] = "Qté prévue"
    ws['D9'] = "Fournisseur"
    ws['E9'] = "Fournisseur"
    ws['F9'] = "Prix unitaire"
    ws['G9'] = "Coût par produit"
    current_row = 10

    for product in query.get_contenu_type(type_):
        for i in range(3):
            ws.cell(row=current_row, column=i+2, value=product[i])
        current_row += 1

    ws_malles = []

    for malle in malles:
        # remove special chars not authorized un sheets
        special_chars = ['[', ']','*', '?', ':', '/', '\\']
        malle = ''.join([i for i in malle if i not in special_chars])   
        wb.create_sheet(malle)
        ws_malles.append(wb.get_sheet_by_name(malle))
        sheet = ws_malles[-1]
        sheet['B2'] = 'Malle :'
        sheet['C2'] = malle
        contenu = query.get_contenu_by_malle(malle)
        header_items = [
            "Désignation Produits :",
            "Qté réelle",
            "Qté prévue",
            "Variation",
            "État"]
        for i, header in enumerate(header_items):
            sheet.cell(column=i+2, row=4, value=header)
            sheet.cell(column=i+2, row=4).border = get_border()
            sheet.cell(column=i+2, row=4).font = styles.Font(bold=True)
        current_row = 5
        for product in contenu:
            logging.debug(product)
            for j, item in enumerate(product):
                sheet.cell(column=j+2, row=current_row, value=item)
                sheet.cell(column=j+2, row=current_row).border = get_border()
            current_row += 1
        
        sheet.row_dimensions[2].font = styles.Font(bold=True, size=16)
        sheet.row_dimensions[2].font = styles.Font(bold=True, size=16)
        sheet.column_dimensions['B'].width = 28

    #sheet['B2'].font = styles.Font(bold=True, size=16)

    wb.save(filename)

def get_border():
    border = styles.Border(
        left=styles.Side(color='000000', style='thin'),
        right=styles.Side(color='000000', style='thin'),
        bottom=styles.Side(color='000000', style='thin'),
        top=styles.Side(color='000000', style='thin'))
    return border

if __name__ == '__main__':
    import sys
    stdout_handler = logging.StreamHandler(sys.stdout)
    stdout_handler.setFormatter(
        logging.Formatter('%(levelname)s::%(module)s:%(lineno)d :: %(message)s'))
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)
    logger.addHandler(stdout_handler)
    
    sys.path.append('./') # must be launch from root app directory
    from db import Query
    q = Query()
    q.connect()
    filename = 'malles_' + type_ \
    + QDate.currentDate().toString('dd-MM-yyyy') + '.xlsx'
    write_file(q, 'Cuisine', filename)
