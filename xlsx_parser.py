#!/usr/bin/python3
# -*- coding: utf-8 -*-

"""
Parse an xlsx (Excel) file to populate the database with malles
"""

from openpyxl import load_workbook
from PyQt5.QtSql import QSqlDatabase, QSqlQuery
from PyQt5.QtCore import QSettings
from db import Query
import re
import logging

def get_id(query):
    req = QSqlQuery(query)
    while req.next():
        return req.value(0)

regex = '^[A-Z]-[A-Z][0-9]{2}$' # regex for location (Adresse col)

def get_col_as_list(col, sheet):
    res = []
    for row in sheet.iter_rows(row_offset=1):
        if isinstance(row[col].value, str) and row[col].value not in res:
            res.append(row[col].value)
    return res

def add_malle_to_db(xlsx_row, lieu_id):
    """ May return False if sheet line seems wrong, 
    error as string if sql insert fail or True if succeed """
    def none_to_str(v):
        if v == None:
            return ''
        else:
            return v
    type_, number, nom, sousclasse, classe, adresse, observation = xlsx_row
    if not type_.value or not number.value:
        return False
    reference = ''.join((type_.value + str(number.value)).split(' '))
    if adresse.value:
        if re.match(regex, adresse.value,flags=re.M):
            section = adresse.value[0]
            slot = adresse.value[2]
            shelf = adresse.value[3:5]
        else:
            section, slot, shelf = '', '', ''
    else:
        section, slot, shelf = '', '', ''
    query = QSqlQuery("INSERT INTO malles(\
    reference, category_id, type_id, lieu_id, section, shelf, slot, observation) \
    VALUES( \
    '" + str(reference) + "', \
    (SELECT id FROM categories WHERE name = '" + str(classe.value) + "'), \
    (SELECT id FROM malles_types WHERE denomination = '" + str(sousclasse.value) + "'), \
    " + str(lieu_id) + ", \
    '" + section + "', \
    '" + shelf + "', \
    '" + slot + "', \
    '" + none_to_str(observation.value) + "')")
    
    if query.lastError().isValid():
        return query.lastError().text()
    else:
        return True

def connect_db():
    db = Query()
    settings = QSettings('Kidivid', 'DiabolikStock')
    connected = db.connect(settings)
    if connected:
        logging.info('database connexion OK')
    return db

def transfert_data(filename, db):
    wb = load_workbook(filename=filename, read_only=True)
    ws = wb.active
    
    # TODO below: must be auto...
    ws.max_row = 525
    ws.max_column = 7 # means 'G' column
    logging.info(ws.calculate_dimension()) # is it good ?
    
    for malle_type in get_col_as_list(3, ws):
        QSqlQuery("INSERT INTO malles_types(denomination) VALUES "\
        + "('" + malle_type + "')")
    for categorie in get_col_as_list(4, ws):
        QSqlQuery("INSERT INTO categories(name) VALUES "\
        + "('" + categorie + "')")
    
    # Location by default
    QSqlQuery("INSERT INTO lieux(nom, ville, cp, numero, rue) VALUES \
    ('Base logistique Ribemont', 'Ribemont-sur-Ancre', 80800, 0, '')")
    lieu_id = get_id("SELECT id FROM lieux WHERE nom = 'Base logistique Ribemont'")

    log = "erreurs:\n"
    for row in ws.iter_rows(row_offset=1):
        result = add_malle_to_db(row, lieu_id=lieu_id)
        if isinstance(result, str):
            log += "Erreur ligne : " + ' '.join([str(i.value) for i in row]) + '\n'
            log += result + '\n'
    return log

if __name__ == '__main__':
    from utils import get_logger
    get_logger()
    db = connect_db()
    transfert_data("liste_malles_2018.xlsx", db)
